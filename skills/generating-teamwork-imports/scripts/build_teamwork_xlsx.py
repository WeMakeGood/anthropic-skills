#!/usr/bin/env python3
"""
Build Teamwork.com import Excel file from audit markdown.

Usage:
    python3 build_teamwork_xlsx.py audit-file.md --output project-import.xlsx

The audit markdown must follow this structure:
    ### Task List Name
    Description: Task list description

    #### Task Name
    - Assignee: Person or Role
    - Due date: YYYY-MM-DD or "Not specified"
    - Estimated time: 30m, 2hr, or "Not specified"
    - Priority: High, Medium, Low (optional)

    ##### Subtasks:
    - [ ] Subtask name (time estimate or "Not specified")
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def parse_audit_markdown(content: str) -> List[Dict]:
    """Parse audit markdown into structured task data."""
    rows = []
    current_tasklist = None
    current_tasklist_desc = None
    current_task = None
    current_task_data = {}
    tasklist_first_row = True
    in_skip_section = False

    lines = content.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # Skip sections that start with h2 headers for non-task content
        if line.startswith('## ') and not line.startswith('### '):
            section_name = line[3:].strip().lower()
            if section_name in ['configuration', 'validation checklist', 'source references']:
                # Save any pending task before entering skip section
                if current_task and current_task_data:
                    row = build_task_row(
                        current_tasklist if tasklist_first_row else None,
                        current_tasklist_desc if tasklist_first_row else None,
                        current_task,
                        current_task_data
                    )
                    rows.append(row)
                    tasklist_first_row = False
                    for subtask in current_task_data.get('subtasks', []):
                        subtask_row = build_subtask_row(subtask, current_task_data)
                        rows.append(subtask_row)
                    current_task = None
                    current_task_data = {}
                in_skip_section = True
                i += 1
                continue
            elif section_name == 'task structure':
                # This is where actual tasks live, exit skip mode
                in_skip_section = False
                i += 1
                continue

        # Task List (h3)
        elif line.startswith('### ') and not line.startswith('#### '):
            tasklist_name = line[4:].strip()

            # Check if we're entering a section to skip
            if tasklist_name.lower() in ['configuration', 'validation checklist', 'source references']:
                # Save any pending task before entering skip section
                if current_task and current_task_data:
                    row = build_task_row(
                        current_tasklist if tasklist_first_row else None,
                        current_tasklist_desc if tasklist_first_row else None,
                        current_task,
                        current_task_data
                    )
                    rows.append(row)
                    tasklist_first_row = False
                    for subtask in current_task_data.get('subtasks', []):
                        subtask_row = build_subtask_row(subtask, current_task_data)
                        rows.append(subtask_row)
                    current_task = None
                    current_task_data = {}
                in_skip_section = True
                i += 1
                continue

            # New valid task list - exit skip mode
            in_skip_section = False

            # Save any pending task before starting new task list
            if current_task and current_task_data:
                row = build_task_row(
                    current_tasklist if tasklist_first_row else None,
                    current_tasklist_desc if tasklist_first_row else None,
                    current_task,
                    current_task_data
                )
                rows.append(row)
                tasklist_first_row = False
                # Add subtasks for the completed task
                for subtask in current_task_data.get('subtasks', []):
                    subtask_row = build_subtask_row(subtask, current_task_data)
                    rows.append(subtask_row)
                current_task = None
                current_task_data = {}
            current_tasklist = tasklist_name
            current_tasklist_desc = None
            tasklist_first_row = True

            # Check next line for description
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line.startswith('Description:'):
                    current_tasklist_desc = next_line[12:].strip()
                    i += 1

        # Skip content in skip sections (validation checklist, etc.)
        elif in_skip_section:
            i += 1
            continue

        # Task (h4)
        elif line.startswith('#### '):
            # Save any pending task first
            if current_task and current_task_data:
                row = build_task_row(
                    current_tasklist if tasklist_first_row else None,
                    current_tasklist_desc if tasklist_first_row else None,
                    current_task,
                    current_task_data
                )
                rows.append(row)
                tasklist_first_row = False
                # Add subtasks for the completed task
                for subtask in current_task_data.get('subtasks', []):
                    subtask_row = build_subtask_row(subtask, current_task_data)
                    rows.append(subtask_row)

            current_task = line[5:].strip()
            current_task_data = {
                'assignee': None,
                'due_date': None,
                'start_date': None,
                'estimated_time': None,
                'priority': None,
                'tags': None,
                'status': 'Active',
                'subtasks': []
            }

        # Task metadata
        elif line.startswith('- Assignee:'):
            if current_task_data is not None:
                current_task_data['assignee'] = parse_metadata_value(line[11:])
        elif line.startswith('- Due date:'):
            if current_task_data is not None:
                current_task_data['due_date'] = parse_date_value(line[11:])
        elif line.startswith('- Start date:'):
            if current_task_data is not None:
                current_task_data['start_date'] = parse_date_value(line[13:])
        elif line.startswith('- Estimated time:'):
            if current_task_data is not None:
                current_task_data['estimated_time'] = parse_time_value(line[17:])
        elif line.startswith('- Priority:'):
            if current_task_data is not None:
                current_task_data['priority'] = parse_metadata_value(line[11:])
        elif line.startswith('- Tags:'):
            if current_task_data is not None:
                current_task_data['tags'] = parse_metadata_value(line[7:])
        elif line.startswith('- Status:'):
            if current_task_data is not None:
                current_task_data['status'] = parse_metadata_value(line[9:])

        # Subtasks section
        elif line.startswith('##### Subtasks:'):
            pass  # Just a header, subtasks follow

        # Subtask item
        elif line.startswith('- [ ]') or line.startswith('- [x]'):
            if current_task_data is not None:
                subtask = parse_subtask(line)
                if subtask:
                    current_task_data['subtasks'].append(subtask)

        i += 1

    # Don't forget the last task
    if current_task and current_task_data:
        row = build_task_row(
            current_tasklist if tasklist_first_row else None,
            current_tasklist_desc if tasklist_first_row else None,
            current_task,
            current_task_data
        )
        rows.append(row)

        # Add subtasks for the last task
        for subtask in current_task_data.get('subtasks', []):
            subtask_row = build_subtask_row(subtask, current_task_data)
            rows.append(subtask_row)

    return rows


def parse_metadata_value(value: str) -> Optional[str]:
    """Parse a metadata value, returning None for 'Not specified'."""
    value = value.strip()
    if value.lower() in ['not specified', 'none', '']:
        return None
    return value


def parse_date_value(value: str) -> Optional[str]:
    """Parse a date value into Teamwork format."""
    value = value.strip()
    if value.lower() in ['not specified', 'none', '']:
        return None

    # Try to parse common date formats
    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%B %d, %Y', '%b %d, %Y']:
        try:
            dt = datetime.strptime(value, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue

    # Return as-is if we can't parse
    return value


def parse_time_value(value: str) -> Optional[str]:
    """Parse a time estimate value into Teamwork format."""
    value = value.strip()
    if value.lower() in ['not specified', 'none', '']:
        return None

    # Already in good format (30m, 2hr, etc.)
    if re.match(r'^\d+[mh]r?$', value.lower()):
        return value.lower()

    # Extract time from parenthetical
    match = re.search(r'(\d+)\s*(hour|hr|h|minute|min|m)', value.lower())
    if match:
        num = match.group(1)
        unit = match.group(2)
        if unit.startswith('h'):
            return f"{num}hr"
        else:
            return f"{num}m"

    return value


def parse_subtask(line: str) -> Optional[Dict]:
    """Parse a subtask line."""
    # Format: - [ ] Subtask name (time estimate)
    # or: - [x] Subtask name (time estimate)

    completed = line.startswith('- [x]')
    content = line[5:].strip() if line.startswith('- [ ]') else line[5:].strip()

    # Extract time estimate from parentheses at end
    time_match = re.search(r'\(([^)]+)\)\s*$', content)
    time_estimate = None
    if time_match:
        time_estimate = parse_time_value(time_match.group(1))
        content = content[:time_match.start()].strip()

    if not content:
        return None

    return {
        'name': content,
        'time_estimate': time_estimate,
        'completed': completed
    }


def build_task_row(tasklist: Optional[str], tasklist_desc: Optional[str],
                   task_name: str, task_data: Dict) -> Dict:
    """Build a row dict for a task."""
    return {
        'TASKLIST': tasklist,
        'TASK': task_name,
        'DESCRIPTION': tasklist_desc,
        'ASSIGN TO': task_data.get('assignee'),
        'START DATE': task_data.get('start_date'),
        'DUE DATE': task_data.get('due_date'),
        'PRIORITY': task_data.get('priority'),
        'ESTIMATED TIME': task_data.get('estimated_time'),
        'TAGS': task_data.get('tags'),
        'STATUS': task_data.get('status', 'Active')
    }


def build_subtask_row(subtask: Dict, parent_task_data: Dict) -> Dict:
    """Build a row dict for a subtask (prefixed with dash)."""
    return {
        'TASKLIST': None,
        'TASK': f"-{subtask['name']}",  # Dash prefix for subtask
        'DESCRIPTION': None,
        'ASSIGN TO': parent_task_data.get('assignee'),  # Inherit from parent
        'START DATE': None,
        'DUE DATE': None,
        'PRIORITY': None,
        'ESTIMATED TIME': subtask.get('time_estimate'),
        'TAGS': None,
        'STATUS': 'Completed' if subtask.get('completed') else 'Active'
    }


def write_excel(rows: List[Dict], output_path: Path) -> None:
    """Write rows to Excel file in Teamwork import format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Headers
    headers = ['TASKLIST', 'TASK', 'DESCRIPTION', 'ASSIGN TO', 'START DATE',
               'DUE DATE', 'PRIORITY', 'ESTIMATED TIME', 'TAGS', 'STATUS']

    # Write headers with bold formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='Build Teamwork.com import Excel file from audit markdown.'
    )
    parser.add_argument('audit_file', type=Path,
                        help='Path to audit markdown file')
    parser.add_argument('--output', '-o', type=Path, required=True,
                        help='Output Excel file path')

    args = parser.parse_args()

    if not args.audit_file.exists():
        print(f"Error: Audit file not found: {args.audit_file}")
        sys.exit(1)

    # Read and parse audit file
    content = args.audit_file.read_text()
    rows = parse_audit_markdown(content)

    if not rows:
        print("Warning: No tasks found in audit file.")
        print("Make sure the file follows the expected markdown structure.")
        sys.exit(1)

    # Write Excel file
    write_excel(rows, args.output)

    # Summary
    task_count = sum(1 for r in rows if not r['TASK'].startswith('-'))
    subtask_count = sum(1 for r in rows if r['TASK'].startswith('-'))
    tasklist_count = sum(1 for r in rows if r['TASKLIST'])

    print(f"Successfully generated: {args.output}")
    print(f"  Task lists: {tasklist_count}")
    print(f"  Tasks: {task_count}")
    print(f"  Subtasks: {subtask_count}")


if __name__ == '__main__':
    main()
